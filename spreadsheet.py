''' def proccess_workbook(filename):
    import openpyxl as xl
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price= cell.value*0.9'''
writing = """ I said "I don't like apples, because I don't like fruit" """
words_list = writing.split(" ")
new_word = ""
for word in words_list:
  not_accepted = ["", '"', '"', '']
  if word in not_accepted:
    pass
  elif word[0] in not_accepted:
    new_word = word.replace(word[0], '')
  elif word[-1] in not_accepted:
    new_word = word.replace(word[-1], '')
  else:
    new_word = word
  words_list_2 = [word]
uniques = []
for word_2 in words_list_2:
  if word_2 not in uniques:
    uniques.append(word_2)
print(uniques)
for word_2 in uniques:
  word_count = words_list_2.count(word_2)
  print(f"the word '{word_2}' is found {word_count} times")